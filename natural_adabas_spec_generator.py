#!/usr/bin/env python3
"""
Natural ADABAS Technical Specification Generator
מחולל אפיון טכני לקוד Natural ADABAS מיינפריים

Usage:
  1. Place this script in the same folder as your .txt Natural files
  2. Run: python natural_adabas_spec_generator.py [--workers N] [--no-synthesis]
  3. Enter your Gemini API key when prompted
  4. Output: natural_adabas_spec.xlsx + knowledge_base.json + analyses.json
     + synthesis.json (consolidated entities + system overview)
     Failed programs are listed in errors.log
  5. After a run: python natural_adabas_spec_generator.py --review [N]
     builds review.html — a smart sample of N programs (default 8) with
     code and analysis side by side, for human quality checking. No API.
  6. python natural_adabas_spec_generator.py --pack [N]
     bundles run conclusions + the selected programs (analysis JSON +
     source code) into a single review_pack.txt, ready to send to an
     AI/expert for deep review. No API.

Resume support:
  Every analyzed program is checkpointed to analyses_checkpoint.jsonl as it
  completes. If the run crashes or is stopped (Ctrl+C), simply run the script
  again — already-analyzed programs are skipped. Delete the checkpoint file
  to force a full re-analysis.

Requirements:
  pip install requests openpyxl
"""

import os
import re
import sys
import json
import time
import hashlib
import threading
import requests
import openpyxl
from pathlib import Path
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────

GEMINI_MODEL  = "gemini-2.5-flash"
GEMINI_BASE   = "https://generativelanguage.googleapis.com/v1beta/models"
MAX_TOKENS    = 65000
TEMPERATURE   = 0.1

# Number of parallel Gemini calls (override with --workers N)
MAX_WORKERS   = 10

# Max characters sent per program to Gemini (to stay within token limits)
MAX_PROGRAM_CHARS = 700_000

# Incremental checkpoint — one JSON line per analyzed program (append-only)
CHECKPOINT_PATH = Path('analyses_checkpoint.jsonl')

# Error log — one line per failed program, for a quick status picture
ERROR_LOG_PATH = Path('errors.log')

# Synthesis checkpoint — consolidated entities (phase 3.5), append-only JSONL
SYNTHESIS_CHECKPOINT_PATH = Path('synthesis_checkpoint.jsonl')

# ──────────────────────────────────────────────────────────────────────────────
# STYLES
# ──────────────────────────────────────────────────────────────────────────────

HDR_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HDR_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=11)
ALT_FILL = PatternFill(start_color="EBF0FA", end_color="EBF0FA", fill_type="solid")
THIN     = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))
WRAP_C   = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_R   = Alignment(horizontal="right",  vertical="top",    wrap_text=True)

def hdr(cell):
    cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, WRAP_C, THIN

def cell_style(cell, alt=False):
    if alt:
        cell.fill = ALT_FILL
    cell.alignment = WRAP_R
    cell.border     = THIN

def set_col_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def freeze_header(ws):
    ws.freeze_panes = "A2"

# ──────────────────────────────────────────────────────────────────────────────
# STATIC PARSER — extract metadata WITHOUT calling AI
# ──────────────────────────────────────────────────────────────────────────────

# Patterns for Natural ADABAS constructs
RE_DEFINE_DATA  = re.compile(r'DEFINE\s+DATA', re.IGNORECASE)
RE_END_DEFINE   = re.compile(r'END-DEFINE', re.IGNORECASE)
# OF is optional in Natural: "01 TASH VIEW NH-TASHLUMIM" is valid syntax
RE_VIEW_OF      = re.compile(r'\bVIEW\s+(?:OF\s+)?([A-Z][A-Z0-9_-]+)', re.IGNORECASE)
RE_CALLNAT      = re.compile(r"CALLNAT\s+['\"]?([A-Z0-9_-]+)['\"]?", re.IGNORECASE)
RE_PERFORM      = re.compile(r'PERFORM\s+([A-Z0-9_-]+)', re.IGNORECASE)
RE_DB_OP        = re.compile(r'\b(READ|FIND|STORE|UPDATE|DELETE|GET\s+SAME|HISTOGRAM)\b', re.IGNORECASE)
RE_PROGRAM_HDR  = re.compile(r'^\*\s*(?:PROGRAM|MODULE|NAME|ROUTINE)\s*[:\-]?\s*(\S+)', re.IGNORECASE)
RE_SUBR_DEF     = re.compile(r'^DEFINE\s+SUBROUTINE\s+([A-Z0-9_-]+)', re.IGNORECASE)
RE_END_SUBR     = re.compile(r'^END-SUBROUTINE', re.IGNORECASE)
RE_IF           = re.compile(r'^\s*IF\b', re.IGNORECASE)
RE_LOOP         = re.compile(r'^\s*(FOR|REPEAT|WHILE|READ|FIND)\b', re.IGNORECASE)

# Executable verbs — used to tell logic members from data-only members
# (LDA/GDA/Copycode). Comment lines start with '*' so they never match.
RE_EXEC_VERB    = re.compile(
    r'^\s*(IF|FOR|REPEAT|DECIDE|COMPUTE|MOVE|ASSIGN|PERFORM|CALLNAT|CALL|INPUT|'
    r'WRITE|DISPLAY|PRINT|FETCH|STORE|UPDATE|DELETE|READ|FIND|HISTOGRAM|ESCAPE)\b',
    re.IGNORECASE | re.MULTILINE)


def classify_program(code):
    """
    'data'  = declarations only (LDA/GDA/Copycode) — analyzed with a short
              field-extraction prompt instead of the full analysis prompt.
    'logic' = everything else (the conservative default).
    """
    if RE_EXEC_VERB.search(code):
        return 'logic'
    if RE_DEFINE_DATA.search(code):
        return 'data'
    return 'logic'


def program_code_hash(code):
    """
    Hash for duplicate detection. The *C** catalog header lines carry
    library/date metadata that differs between exports of identical source,
    so they are stripped before hashing, as is trailing whitespace.
    """
    lines = [l.rstrip() for l in code.splitlines() if not RE_CATALOG_SEP.match(l)]
    return hashlib.sha1('\n'.join(lines).strip().encode('utf-8')).hexdigest()


RE_END_PROG_SEMI     = re.compile(r'^\s*END\s*;?\s*$', re.IGNORECASE)
RE_PROG_NAME_CMT     = re.compile(r'^\*{1,2}\s*(?:PROGRAM|PROG|MODULE|NAME|ROUTINE)\s*[:\-]?\s*([A-Z0-9_-]+)', re.IGNORECASE)
RE_DEFINE_DATA_START = re.compile(r'^\s*DEFINE\s+DATA\b', re.IGNORECASE)
# *C** lines in mainframe exports: *C**   <LIBRARY>   <PROGNAME>   <metadata...>
# They mark the start of each cataloged member. Match any line starting with *C**
RE_CATALOG_SEP       = re.compile(r'^\*C\*\*')


def _extract_name_from_segment(lines, stem, index):
    """
    Extract program name from a segment's *C** header line.
    Format: *C**   <LIBRARY>   <PROGNAME>   <metadata>
    The program name is the 3rd whitespace-separated token (index 2).
    Falls back to comment headers or stem_PROGn.
    """
    if lines:
        first = lines[0].strip()
        if first.startswith('*C**'):
            tokens = first.split()
            # tokens[0]='*C**', tokens[1]=library, tokens[2]=progname
            if len(tokens) >= 3:
                return tokens[2].upper()
            elif len(tokens) == 2:
                return tokens[1].upper()
    # Fallback: look for comment-style name in first 15 lines
    for line in lines[:15]:
        m = RE_PROG_NAME_CMT.match(line.strip())
        if m:
            return m.group(1).upper()
    return f"{stem}_PROG{index + 1}"


def split_into_programs(content, filename):
    """
    Split a Natural ADABAS file into individual programs.

    Priority order:
      A. *C** separator lines  (mainframe export convention)
      B. Standalone END lines  (Natural language standard)
      C. Multiple DEFINE DATA  (heuristic)
      D. Whole file as one unit (fallback)
    """
    stem  = Path(filename).stem
    lines = content.splitlines()

    # ── Strategy A: *C** separator lines ────────────────────────────────────
    sep_indices = [i for i, l in enumerate(lines) if RE_CATALOG_SEP.match(l)]
    if sep_indices:
        segments = []
        boundaries = sep_indices + [len(lines)]
        for idx, start in enumerate(sep_indices):
            end     = boundaries[idx + 1]
            seg     = lines[start:end]
            name    = _extract_name_from_segment(seg, stem, idx)
            segments.append({'name': name, 'code': '\n'.join(seg)})
        if segments:
            return segments

    # ── Strategy B: standalone END lines ────────────────────────────────────
    segments = []
    current  = []

    for line in lines:
        current.append(line)
        if RE_END_PROG_SEMI.match(line):
            non_empty = [l for l in current if l.strip() and not l.strip().startswith('*')]
            if len(non_empty) >= 3:
                segments.append(current[:])
            current = []

    if current:
        non_empty = [l for l in current if l.strip() and not l.strip().startswith('*')]
        if len(non_empty) >= 3:
            segments.append(current)

    # If we found more than 1 segment, use them
    if len(segments) > 1:
        programs = []
        for i, seg_lines in enumerate(segments):
            name = _extract_name_from_segment(seg_lines, stem, i)
            programs.append({'name': name, 'code': '\n'.join(seg_lines)})
        return programs

    # ── Strategy C: split on DEFINE DATA blocks ──────────────────────────────
    split_points = [i for i, l in enumerate(lines) if RE_DEFINE_DATA_START.match(l)]

    if len(split_points) > 1:
        programs = []
        for idx, start in enumerate(split_points):
            end = split_points[idx + 1] if idx + 1 < len(split_points) else len(lines)
            seg = lines[start:end]
            name = _extract_name_from_segment(seg, stem, idx)
            programs.append({'name': name, 'code': '\n'.join(seg)})
        return programs

    # ── Fallback: whole file as one program ──────────────────────────────────
    return [{'name': stem, 'code': content}]


def debug_file(filepath):
    """
    Print diagnostic info to help tune the program splitter.
    Run with:  python natural_adabas_spec_generator.py --debug <file.txt>
    """
    with open(filepath, 'rb') as f:
        raw = f.read()

    # Detect line ending style
    crlf = raw.count(b'\r\n')
    lf   = raw.count(b'\n') - crlf
    print(f"\nFile: {filepath}  ({len(raw):,} bytes)")
    print(f"Line endings: CRLF={crlf}  LF={lf}")

    lines = raw.decode('utf-8', errors='replace').splitlines()
    print(f"Total lines: {len(lines)}")

    # Show first 10 lines with hex
    print("\n── First 10 lines (repr) ──────────────────────────")
    for i, l in enumerate(lines[:10]):
        print(f"  {i+1:4d}: {repr(l)}")

    # Find candidate separator lines (lines with only * and C chars)
    print("\n── Candidate separator lines (contain *C) ─────────")
    found = 0
    for i, l in enumerate(lines):
        stripped = l.strip()
        if stripped.startswith('*C') or stripped.startswith('* C'):
            print(f"  line {i+1:5d}: {repr(l)}")
            found += 1
            if found >= 20:
                print("  ... (showing first 20)")
                break
    if found == 0:
        print("  (none found — looking for any line starting with *)")
        for i, l in enumerate(lines):
            if l.strip().startswith('*') and len(l.strip()) < 20:
                print(f"  line {i+1:5d}: {repr(l)}")
                found += 1
                if found >= 20:
                    break

    # Show DEFINE DATA occurrences
    dd = [i+1 for i, l in enumerate(lines) if re.match(r'^\s*DEFINE\s+DATA\b', l, re.IGNORECASE)]
    print(f"\n── DEFINE DATA at lines: {dd}")

    # Show END occurrences (standalone)
    ends = [i+1 for i, l in enumerate(lines) if re.match(r'^\s*END\s*\r?$', l, re.IGNORECASE)]
    print(f"── Standalone END at lines: {ends[:20]}")
    sys.exit(0)


# Some mainframe exports prefix every SOURCE line with *S** (catalog lines
# are *C**, metadata *D..). Without de-prefixing, all code looks like
# comments: programs get classified as data-only and regexes match nothing.
RE_SOURCE_PREFIX = re.compile(r'^\*S\*\*')


def strip_source_prefix(code):
    """Remove the *S** source-line prefix when it dominates the segment.
    A prefixed comment (*S***...) correctly becomes a plain * comment."""
    lines     = code.splitlines()
    non_empty = [l for l in lines if l.strip()]
    n_pref    = sum(1 for l in non_empty if RE_SOURCE_PREFIX.match(l))
    if n_pref >= 3 and n_pref >= 0.3 * max(1, len(non_empty)):
        return '\n'.join(RE_SOURCE_PREFIX.sub('', l) for l in lines)
    return code


def parse_file(filepath):
    """Static parse of one .txt Natural file — no AI required."""
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    filename = os.path.basename(filepath)

    # Split first (the *C** separators must stay intact), then de-prefix
    # each program's source so the real code is visible to all regexes
    programs = split_into_programs(content, filename)
    for prog in programs:
        prog['code']     = strip_source_prefix(prog['code'])
        prog['callnats'] = sorted({m.group(1).upper()
                                   for m in RE_CALLNAT.finditer(prog['code'])})
        prog['ptype']    = classify_program(prog['code'])
        prog['hash']     = program_code_hash(prog['code'])
        prog['ddms']     = sorted({m.group(1).upper()
                                   for m in RE_VIEW_OF.finditer(prog['code'])})

    # File-level scan over the de-prefixed source
    ddms      = set()
    callnats  = set()
    performs  = set()
    db_ops    = []
    clean_lines = '\n'.join(p['code'] for p in programs).splitlines()
    for i, line in enumerate(clean_lines, 1):
        s = line.strip()
        for m in RE_VIEW_OF.finditer(s):
            ddms.add(m.group(1).upper())
        for m in RE_CALLNAT.finditer(s):
            callnats.add(m.group(1).upper())
        for m in RE_PERFORM.finditer(s):
            performs.add(m.group(1).upper())
        for m in RE_DB_OP.finditer(s):
            db_ops.append({'op': m.group(1).upper(), 'line': i, 'context': s[:80]})

    return {
        'filename' : filename,
        'content'  : content,
        'programs' : programs,
        'ddms'     : sorted(ddms),
        'callnats' : sorted(callnats),
        'performs' : sorted(performs),
        'db_ops'   : db_ops,
        'line_count': len(content.splitlines()),
    }


def build_knowledge_base(files_data):
    """Build cross-reference map across all files."""
    ddm_to_files  = defaultdict(list)
    call_graph    = defaultdict(list)  # file → [called programs]
    all_ddms      = set()
    program_index = {}  # program_name → filename

    for fd in files_data:
        fname = fd['filename']
        for ddm in fd['ddms']:
            ddm_to_files[ddm].append(fname)
            all_ddms.add(ddm)
        for cn in fd['callnats']:
            call_graph[fname].append(cn)
        for prog in fd['programs']:
            program_index[prog['name'].upper()] = fname

    return {
        'ddm_to_files' : dict(ddm_to_files),
        'call_graph'   : dict(call_graph),
        'all_ddms'     : sorted(all_ddms),
        'program_index': program_index,
    }


def _norm_name(s):
    """Normalize a technical name for fuzzy matching (strip punctuation/case)."""
    return re.sub(r'[^A-Z0-9]', '', (s or '').upper())


def build_validation(files_data, analyses):
    """
    Cross-check the static regex findings against what the AI reported,
    per program. No API calls. Returns rows of:
      (filename, program, kind, name, finding)

    Severity tiers for AI claims with no anchor in the program's own code:
      - the name exists elsewhere (same file / anywhere in the system, e.g.
        defined in an external LDA) → "שיוך שגוי" (attribution issue, mild)
      - the name exists nowhere in the system → "חשד להזיה" (real concern)
    Code findings the AI didn't report are misses.
    """
    by_key = {(a.get('filename', ''), a.get('program_name', '')): a for a in analyses}

    # Global static inventories — for telling attribution issues from inventions
    all_ddms_n = {_norm_name(d) for fd in files_data for d in fd['ddms']}
    all_targets_n = set()                       # known program names + call targets
    for fd in files_data:
        for p in fd['programs']:
            all_targets_n.add(_norm_name(p['name']))
            all_targets_n.update(_norm_name(c) for c in p.get('callnats', []))

    rows = []
    for fd in files_data:
        file_ddms_n = {_norm_name(d) for d in fd['ddms']}
        for prog in fd['programs']:
            a = by_key.get((fd['filename'], prog['name']))
            if not a or 'error' in a:
                continue
            code           = prog['code']
            static_ddms    = {m.group(1).upper() for m in RE_VIEW_OF.finditer(code)}
            static_ddms_n  = {_norm_name(d) for d in static_ddms}
            static_calls   = {m.group(1).upper() for m in RE_CALLNAT.finditer(code)}
            static_calls_n = {_norm_name(c) for c in static_calls}
            ai_entities    = {(e.get('name') or '').strip().upper()
                              for e in a.get('entities', [])} - {''}
            ai_callnats    = {(d.get('name') or '').strip().upper()
                              for d in a.get('dependencies', [])
                              if (d.get('type') or '').upper() == 'CALLNAT'} - {''}

            for name in sorted(ai_entities):
                n = _norm_name(name)
                if n in static_ddms_n:
                    continue
                if n in file_ddms_n:
                    rows.append((fd['filename'], prog['name'], 'ישות', name,
                                 'שיוך שגוי — הישות קיימת בקובץ אך אין לה VIEW OF בתוכנית זו (ייתכן LDA חיצוני)'))
                elif n in all_ddms_n:
                    rows.append((fd['filename'], prog['name'], 'ישות', name,
                                 'שיוך שגוי — הישות קיימת במערכת אך לא בתוכנית זו'))
                else:
                    rows.append((fd['filename'], prog['name'], 'ישות', name,
                                 'הישות לא נמצאה בשום מקום במערכת — חשד להזיה'))
            for name in sorted(static_ddms):
                if _norm_name(name) not in {_norm_name(e) for e in ai_entities}:
                    rows.append((fd['filename'], prog['name'], 'ישות', name,
                                 'נמצא VIEW OF בקוד אך ה-AI לא דיווח על הישות — פספוס'))

            for name in sorted(ai_callnats):
                n = _norm_name(name)
                if n in static_calls_n:
                    continue
                if n in all_targets_n:
                    rows.append((fd['filename'], prog['name'], 'CALLNAT', name,
                                 'שיוך שגוי — התוכנית קיימת במערכת אך אין CALLNAT אליה בקוד תוכנית זו'))
                else:
                    rows.append((fd['filename'], prog['name'], 'CALLNAT', name,
                                 'תוכנית היעד לא נמצאה בשום מקום במערכת — חשד להזיה'))
            for name in sorted(static_calls):
                if _norm_name(name) not in {_norm_name(c) for c in ai_callnats}:
                    rows.append((fd['filename'], prog['name'], 'CALLNAT', name,
                                 'נמצא CALLNAT בקוד אך ה-AI לא דיווח עליו — פספוס'))
    return rows


# ──────────────────────────────────────────────────────────────────────────────
# REVIEW MODE (--review) — human quality-check package, no API calls
# ──────────────────────────────────────────────────────────────────────────────

def select_review_sample(files_data, analyses, validation_rows, n=8):
    """
    Pick a smart (not random) sample for human review:
    the most complex programs, the most-called ones, programs with many
    validation gaps, a couple of medium ones, and one data member.
    Returns [(analysis, reason), ...].
    """
    ok = [a for a in analyses if 'error' not in a]

    gaps = defaultdict(int)
    for r in validation_rows:
        gaps[(r[0], r[1])] += 1

    callers = defaultdict(int)          # program name → how many programs call it
    for fd in files_data:
        for p in fd['programs']:
            for c in p.get('callnats', []):
                callers[c] += 1

    def key(a):       return (a.get('filename', ''), a.get('program_name', ''))
    def n_rules(a):   return sum(len(w.get('business_rules', [])) for w in a.get('workflows', []))
    def richness(a):  return len(a.get('workflows', [])) + n_rules(a)
    def n_callers(a): return callers.get((a.get('program_name') or '').upper(), 0)

    sample, seen = [], set()
    def add(a, reason):
        if key(a) not in seen and len(sample) < n:
            seen.add(key(a))
            sample.append((a, reason))

    for a in sorted([a for a in ok if a.get('complexity') == 'מורכב'],
                    key=richness, reverse=True)[:3]:
        add(a, f"most complex ({len(a.get('workflows', []))} flows, {n_rules(a)} rules)")
    for a in sorted(ok, key=n_callers, reverse=True)[:2]:
        if n_callers(a) > 0:
            add(a, f"central in call graph ({n_callers(a)} callers)")
    for a in sorted(ok, key=lambda x: gaps.get(key(x), 0), reverse=True)[:2]:
        if gaps.get(key(a), 0) >= 2:
            add(a, f"{gaps[key(a)]} validation gaps")
    for a in [a for a in ok if a.get('complexity') == 'בינוני'][:2]:
        add(a, "medium complexity")
    for a in [a for a in ok if a.get('program_type') == 'data'][:1]:
        add(a, "data-only member (short prompt)")
    # top up with the richest programs if the sample is still small
    for a in sorted(ok, key=richness, reverse=True):
        add(a, "sample top-up")
    return sample[:n]


def _render_analysis_html(a, esc):
    parts = [f"<p><b>מטרה עסקית:</b> {esc(a.get('business_purpose', ''))}</p>",
             f"<p><b>מורכבות:</b> {esc(a.get('complexity', ''))}"
             + (f" &middot; <b>שוכפל מ:</b> {esc(a['duplicate_of'])}" if a.get('duplicate_of') else "")
             + "</p>"]
    if a.get('entities'):
        parts.append("<h4>ישויות</h4><ul>")
        for e in a['entities']:
            fields = e.get('fields', [])
            f_txt = ", ".join(f"{esc(f.get('name', ''))} ({esc(f.get('type', ''))})"
                              for f in fields[:12])
            if len(fields) > 12:
                f_txt += f" ועוד {len(fields) - 12}"
            parts.append(f"<li><b>{esc(e.get('name', ''))}</b> — {esc(e.get('description', ''))}"
                         f"<div class='fields'>{len(fields)} שדות: {f_txt}</div></li>")
        parts.append("</ul>")
    if a.get('workflows'):
        parts.append("<h4>זרימות עבודה</h4>")
        for w in a['workflows']:
            parts.append(f"<div class='wf'><b>{esc(w.get('name', ''))}</b> "
                         f"<span class='trig'>({esc(w.get('trigger', ''))} — "
                         f"{esc(w.get('trigger_details', ''))})</span><ol>")
            parts += [f"<li>{esc(s)}</li>" for s in w.get('steps', [])]
            parts.append("</ol>")
            if w.get('business_rules'):
                parts.append("<ul class='rules'>")
                parts += [f"<li><b>{esc(r.get('name', ''))}:</b> "
                          f"<code>{esc(r.get('logic', ''))}</code></li>"
                          for r in w['business_rules']]
                parts.append("</ul>")
            parts.append("</div>")
    if a.get('permissions'):
        parts.append("<h4>הרשאות</h4><ul>")
        parts += [f"<li><b>{esc(p.get('role', ''))}</b> על {esc(p.get('entity', ''))} — "
                  f"קריאה:{esc(p.get('read', ''))} עדכון:{esc(p.get('update', ''))} "
                  f"מחיקה:{esc(p.get('delete', ''))}</li>" for p in a['permissions']]
        parts.append("</ul>")
    if a.get('open_questions'):
        parts.append("<h4>שאלות פתוחות</h4><ul>"
                     + "".join(f"<li>{esc(q)}</li>" for q in a['open_questions']) + "</ul>")
    return "\n".join(parts)


REVIEW_CSS = """
body{font-family:Arial,'Segoe UI',sans-serif;margin:0;background:#f4f6fa;color:#1a2233}
header{background:#1e3a5f;color:#fff;padding:18px 28px}
header h1{margin:0 0 6px;font-size:22px} header p{margin:0;opacity:.85;font-size:13px}
.stats{display:flex;gap:14px;padding:16px 28px;flex-wrap:wrap}
.card{background:#fff;border-radius:8px;padding:12px 18px;box-shadow:0 1px 3px rgba(0,0,0,.12);min-width:130px}
.card .num{font-size:24px;font-weight:bold} .card.warn .num{color:#c0392b}
.checklist{margin:0 28px 10px;background:#fff8e1;border-right:4px solid #f0a818;border-radius:6px;padding:10px 16px;font-size:14px}
section{background:#fff;margin:18px 28px;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.12);overflow:hidden}
section>h2{background:#2e5990;color:#fff;margin:0;padding:10px 16px;font-size:16px}
section>h2 .badge{background:#f0a818;color:#1a2233;border-radius:10px;padding:2px 10px;font-size:12px;margin-inline-start:10px}
.cols{display:grid;grid-template-columns:1fr 1fr;gap:0}
.analysis{padding:14px 18px;border-inline-end:1px solid #e2e8f2;font-size:14px}
.analysis h4{margin:14px 0 6px;color:#1e3a5f}
.fields{color:#5a6a85;font-size:12px;margin:3px 0 6px}
.wf{margin:0 0 12px;padding:8px 10px;background:#f4f7fc;border-radius:6px}
.wf .trig{color:#5a6a85;font-size:12px} .wf ol{margin:6px 0;padding-inline-start:20px}
.rules{background:#fff;border-radius:4px;margin:4px 0;padding:6px 22px;font-size:13px}
.gaps{margin:10px 18px;padding:8px 12px;background:#fdecea;border-radius:6px;font-size:13px}
pre{margin:0;padding:14px;background:#10141c;color:#d7e0ee;font-size:11.5px;line-height:1.5;overflow:auto;max-height:780px;direction:ltr;text-align:left}
@media print{pre{max-height:none}}
"""


def generate_review_html(sample, code_map, validation_rows, stats, out_path):
    from html import escape as esc
    gaps_by_prog = defaultdict(list)
    for r in validation_rows:
        gaps_by_prog[(r[0], r[1])].append(r)

    top_gap_files = defaultdict(int)
    for r in validation_rows:
        top_gap_files[r[0]] += 1
    top_files_txt = ", ".join(f"{esc(f)} ({c})" for f, c in
                              sorted(top_gap_files.items(), key=lambda x: -x[1])[:5])

    h = [f"<!DOCTYPE html><html dir='rtl' lang='he'><head><meta charset='utf-8'>"
         f"<title>סקירת איכות — אפיון Natural ADABAS</title><style>{REVIEW_CSS}</style></head><body>",
         f"<header><h1>סקירת איכות — אפיון Natural ADABAS</h1>"
         f"<p>נוצר {time.strftime('%Y-%m-%d %H:%M')} · דגימה של {len(sample)} תוכניות מתוך {stats['total']}</p></header>",
         "<div class='stats'>",
         f"<div class='card'><div class='num'>{stats['total']}</div>תוכניות נותחו</div>",
         f"<div class='card{' warn' if stats['errors'] else ''}'><div class='num'>{stats['errors']}</div>שגיאות</div>",
         f"<div class='card{' warn' if stats['halluc'] else ''}'><div class='num'>{stats['halluc']}</div>חשדות הזיה</div>",
         f"<div class='card'><div class='num'>{stats.get('attrib', 0)}</div>שיוכים שגויים</div>",
         f"<div class='card{' warn' if stats['misses'] else ''}'><div class='num'>{stats['misses']}</div>פספוסים</div>",
         "</div>",
         "<div class='checklist'><b>מקרא:</b> חשד הזיה = שם שלא קיים בשום מקום במערכת (חמור). "
         "שיוך שגוי = ישות/תוכנית אמיתית שיוחסה לתוכנית הלא נכונה — לרוב הגדרה ב-LDA חיצוני "
         "או הקשר קובץ; כמעט לא פוגע בגיליונות המאוחדים. פספוס = נמצא בקוד ולא דווח.</div>"]
    if top_files_txt:
        h.append(f"<div class='checklist'>ריכוזי פערים לפי קובץ: {top_files_txt}</div>")
    h.append("<div class='checklist'><b>לכל תוכנית בדגימה, בדוק 4 שאלות:</b> "
             "1) המטרה העסקית נכונה? &middot; 2) הזרימות תואמות את הקוד (PF-keys, הסתעפויות)? &middot; "
             "3) חוקי העסק קיימים בקוד ובכיוון הנכון (&lt; לעומת &gt;)? &middot; "
             "4) השדות והטיפוסים תואמים ל-DEFINE DATA?</div>")

    for a, reason in sample:
        k = (a.get('filename', ''), a.get('program_name', ''))
        code = code_map.get(k, '(הקוד לא נמצא)')
        if len(code) > 30000:
            code = code[:30000] + "\n\n... (קוצר לתצוגה — הקוד המלא בקובץ המקור)"
        h.append(f"<section><h2>{esc(k[1])} <small>({esc(k[0])})</small>"
                 f"<span class='badge'>{esc(reason)}</span></h2>")
        prog_gaps = gaps_by_prog.get(k)
        if prog_gaps:
            h.append("<div class='gaps'><b>פערי ולידציה בתוכנית זו:</b><br>"
                     + "<br>".join(f"{esc(r[2])} {esc(r[3])} — {esc(r[4])}" for r in prog_gaps)
                     + "</div>")
        h.append(f"<div class='cols'><div class='analysis'>{_render_analysis_html(a, esc)}</div>"
                 f"<pre>{esc(code)}</pre></div></section>")

    h.append("</body></html>")
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(h))


def _load_review_inputs():
    """Shared loader for --review/--pack: analyses + parsed sources + stats."""
    analyses_path = Path('analyses.json')
    if analyses_path.exists():
        with open(analyses_path, 'r', encoding='utf-8') as f:
            analyses = json.load(f)
    else:
        ckpt = load_checkpoint(CHECKPOINT_PATH)
        if not ckpt:
            print("No analyses.json or checkpoint found — run the analysis first.")
            sys.exit(1)
        analyses = list(ckpt.values())

    txt_files  = sorted(Path('.').glob('*.txt'))
    files_data = [parse_file(fp) for fp in txt_files]
    code_map   = {(fd['filename'], p['name']): p['code']
                  for fd in files_data for p in fd['programs']}

    validation_rows = build_validation(files_data, analyses)
    stats = {
        'total' : len([a for a in analyses if 'error' not in a]),
        'errors': len([a for a in analyses if 'error' in a]),
        'halluc': len([r for r in validation_rows if 'חשד להזיה' in r[4]]),
        'attrib': len([r for r in validation_rows if 'שיוך שגוי' in r[4]]),
        'misses': len([r for r in validation_rows if 'פספוס' in r[4]]),
    }
    return analyses, files_data, code_map, validation_rows, stats


def _print_review_summary(stats, sample):
    print("=" * 62)
    print("  Quality review — summary")
    print("=" * 62)
    print(f"  Programs analyzed: {stats['total']}   errors: {stats['errors']}")
    print(f"  Hallucination suspicions: {stats['halluc']}   "
          f"misattributions: {stats['attrib']}   misses: {stats['misses']}")
    print(f"\n  Selected sample ({len(sample)} programs):")
    for a, reason in sample:
        print(f"    - {a.get('program_name', ''):20s} ({a.get('filename', '')})  — {reason}")


def run_review(sample_size=8):
    """--review entry point: build review.html from existing outputs. No API."""
    analyses, files_data, code_map, validation_rows, stats = _load_review_inputs()
    sample = select_review_sample(files_data, analyses, validation_rows, sample_size)

    out_path = Path('review.html')
    generate_review_html(sample, code_map, validation_rows, stats, out_path)

    _print_review_summary(stats, sample)
    print(f"\n  {out_path} created — open in a browser and review program by program")


PACK_CODE_CHAR_CAP = 40_000   # per program, keeps the pack sendable


def run_pack(sample_size=8):
    """
    --pack entry point: bundle run conclusions + the selected programs
    (full analysis JSON + full source code) into ONE text file,
    review_pack.txt, ready to hand to an AI/expert for deep review. No API.
    """
    analyses, files_data, code_map, validation_rows, stats = _load_review_inputs()
    sample = select_review_sample(files_data, analyses, validation_rows, sample_size)

    gaps_by_prog = defaultdict(list)
    for r in validation_rows:
        gaps_by_prog[(r[0], r[1])].append(r)
    gap_files = defaultdict(int)
    for r in validation_rows:
        gap_files[r[0]] += 1

    err_entries = [a for a in analyses if 'error' in a]

    L = []
    L.append("=" * 78)
    L.append("REVIEW PACK — Natural ADABAS spec generator")
    L.append(f"Generated: {time.strftime('%Y-%m-%d %H:%M')}")
    L.append("Purpose: deep quality review of AI-generated analyses against source code.")
    L.append("For each program below: check business purpose, workflow coverage of all")
    L.append("code paths (incl. subroutines), rule correctness INCLUDING direction")
    L.append("(< vs >, before/after dates), and fields vs DEFINE DATA.")
    L.append("=" * 78)
    L.append("")
    L.append("== RUN CONCLUSIONS ==")
    L.append(f"Programs analyzed OK : {stats['total']}")
    L.append(f"Failed programs      : {stats['errors']}")
    L.append(f"Hallucination susp.  : {stats['halluc']}  (name found nowhere in system)")
    L.append(f"Misattributions      : {stats['attrib']}  (real name, wrong program — usually external LDA)")
    L.append(f"Misses               : {stats['misses']}  (in code, not reported)")
    if gap_files:
        top = sorted(gap_files.items(), key=lambda x: -x[1])[:8]
        L.append("Gap concentration by file: " + ", ".join(f"{f}({c})" for f, c in top))
    if err_entries:
        L.append("")
        L.append(f"Failed programs ({min(len(err_entries), 15)} of {len(err_entries)}):")
        for a in err_entries[:15]:
            L.append(f"  - {a.get('filename', '')} / {a.get('program_name', '')}: "
                     f"{a.get('error', '')}")
    L.append("")
    L.append(f"== SELECTED PROGRAMS ({len(sample)}) ==")
    for a, reason in sample:
        L.append(f"  - {a.get('program_name', '')} ({a.get('filename', '')}) — {reason}")
    L.append("")

    for i, (a, reason) in enumerate(sample, 1):
        k    = (a.get('filename', ''), a.get('program_name', ''))
        code = code_map.get(k, '(source code not found)')
        if len(code) > PACK_CODE_CHAR_CAP:
            code = code[:PACK_CODE_CHAR_CAP] + "\n... (truncated for packing)"
        L.append("#" * 78)
        L.append(f"# PROGRAM {i}/{len(sample)}: {k[1]}   (file: {k[0]})")
        L.append(f"# Selected because: {reason}")
        L.append("#" * 78)
        prog_gaps = gaps_by_prog.get(k)
        if prog_gaps:
            L.append("-- VALIDATION GAPS FOR THIS PROGRAM --")
            for r in prog_gaps:
                L.append(f"  [{r[2]}] {r[3]}: {r[4]}")
        L.append("-- AI ANALYSIS (JSON) --")
        L.append(json.dumps(a, ensure_ascii=False, indent=2))
        L.append("-- SOURCE CODE --")
        L.append(code)
        L.append("")

    out_path = Path('review_pack.txt')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(L))

    _print_review_summary(stats, sample)
    size_kb = out_path.stat().st_size / 1024
    print(f"\n  {out_path} created ({size_kb:.0f} KB) — send this single file for deep review")


# ──────────────────────────────────────────────────────────────────────────────
# GEMINI API
# ──────────────────────────────────────────────────────────────────────────────

class GeminiTruncated(Exception):
    """The response hit MAX_TOKENS and was cut off mid-output."""
    def __init__(self, partial_text):
        super().__init__("Response truncated (MAX_TOKENS)")
        self.partial_text = partial_text


def call_gemini(api_key, prompt, attempt=0):
    """Call Gemini REST API with exponential-backoff retry."""
    url = f"{GEMINI_BASE}/{GEMINI_MODEL}:generateContent?key={api_key}"
    body = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {
            "maxOutputTokens": MAX_TOKENS,
            "temperature": TEMPERATURE,
            # JSON mode — forces syntactically valid JSON output
            "responseMimeType": "application/json",
        },
    }
    try:
        resp = requests.post(url, json=body, timeout=360)
        if resp.status_code in (429, 503) and attempt < 4:
            wait = 30 * (2 ** attempt)  # 30s, 60s, 120s, 240s
            print(f"    Rate limit ({resp.status_code}) — waiting {wait}s...", flush=True)
            time.sleep(wait)
            return call_gemini(api_key, prompt, attempt + 1)
        resp.raise_for_status()
        data      = resp.json()
        candidate = (data.get("candidates") or [{}])[0]
        parts     = candidate.get("content", {}).get("parts", [])
        text      = "".join(p.get("text", "") for p in parts)
        if candidate.get("finishReason") == "MAX_TOKENS":
            raise GeminiTruncated(text)
        return text
    except requests.Timeout:
        if attempt < 3:
            wait = 30 * (2 ** attempt)  # 30s, 60s, 120s
            print(f"    Timeout — waiting {wait}s before retry...", flush=True)
            time.sleep(wait)
            return call_gemini(api_key, prompt, attempt + 1)
        raise
    except requests.RequestException as e:
        if attempt < 4:
            wait = 15 * (2 ** attempt)  # 15s, 30s, 60s, 120s
            print(f"    Network error: {e} — waiting {wait}s...", flush=True)
            time.sleep(wait)
            return call_gemini(api_key, prompt, attempt + 1)
        raise


def strip_json_fences(text):
    """Remove ```json ... ``` markdown fences if present."""
    text = text.strip()
    text = re.sub(r'^```(?:json)?\s*', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\s*```$',          '', text)
    return text.strip()


def parse_json_response(raw):
    """
    Parse a Gemini response into a dict. If the model appended text after
    the JSON object ("Extra data" errors), salvage the first complete object.
    """
    cleaned = strip_json_fences(raw)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find('{')
        if start == -1:
            raise
        obj, _ = json.JSONDecoder().raw_decode(cleaned[start:])
        return obj


def log_error(path, filename, prog_name, message):
    """Append one line per failed program — a quick status picture."""
    ts = time.strftime('%Y-%m-%d %H:%M:%S')
    with open(path, 'a', encoding='utf-8') as f:
        f.write(f"{ts}  {filename} › {prog_name}  {message}\n")


CONCISE_SUFFIX = """

חשוב: התשובה הקודמת נקטעה באמצע כי הייתה ארוכה מדי.
קצר משמעותית: עד 5 הזרימות המרכזיות בלבד, עד 10 שדות חשובים לכל ישות,
תיאורים של משפט אחד. שמור על מבנה ה-JSON המבוקש."""


def call_and_parse(api_key, prompt):
    """
    Call Gemini and parse the JSON answer, recovering from the two common
    failure modes:
      - MAX_TOKENS truncation → one retry asking for a concise answer
      - malformed JSON        → one retry
    Returns (analysis, None) on success or (None, error_message) on failure.
    Network errors propagate to the caller.
    """
    truncated = False
    json_err  = None
    current   = prompt
    for _ in range(3):
        try:
            raw = call_gemini(api_key, current)
        except GeminiTruncated:
            if truncated:
                return None, "Truncated (MAX_TOKENS) even after concise retry"
            truncated = True
            current   = prompt + CONCISE_SUFFIX
            continue
        try:
            return parse_json_response(raw), None
        except json.JSONDecodeError as e:
            if json_err is not None:
                return None, f"JSON parse: {e}"
            json_err = e
    if json_err is not None:
        return None, f"JSON parse: {json_err}"
    return None, "No valid response received"


# ──────────────────────────────────────────────────────────────────────────────
# CHECKPOINT (resume support)
# ──────────────────────────────────────────────────────────────────────────────

def analysis_key(a):
    return (a.get('filename', ''), a.get('program_name', ''))


def load_checkpoint(path, key_fn=analysis_key):
    """
    Load prior entries from a JSONL checkpoint file, keyed by key_fn.
    If a key appears more than once (e.g. a failed attempt followed by a
    successful retry), the last entry wins. A truncated final line
    (killed mid-write) is skipped.
    """
    results = {}
    if not path.exists():
        return results
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                a = json.loads(line)
            except json.JSONDecodeError:
                continue
            results[key_fn(a)] = a
    return results


def append_checkpoint(path, analysis):
    # If a previous run was killed mid-write, the file may end with a
    # truncated line lacking '\n' — terminate it first so the new entry
    # lands on its own line and stays parseable.
    needs_nl = path.exists() and path.stat().st_size > 0
    if needs_nl:
        with open(path, 'rb') as f:
            f.seek(-1, 2)
            needs_nl = f.read(1) != b'\n'
    with open(path, 'ab') as f:
        if needs_nl:
            f.write(b'\n')
        f.write((json.dumps(analysis, ensure_ascii=False) + '\n').encode('utf-8'))


# ──────────────────────────────────────────────────────────────────────────────
# PROMPTS
# ──────────────────────────────────────────────────────────────────────────────

SYSTEM_CONTEXT = """אתה מנתח קוד Natural ADABAS מיינפריים בכיר.
ענה תמיד בעברית, מלבד שמות טכניים (DDM, שדות, שמות תוכניות).

רקע:
- Natural = שפת תכנות מיינפריים של Software AG
- ADABAS = בסיס נתונים מיינפריים של Software AG
- DDM (Data Definition Module) = הגדרת טבלה / ישות
- CALLNAT = קריאה לתת-תוכנית חיצונית
- PERFORM = קריאה לתת-שגרה באותו קובץ
- DEFINE DATA / VIEW OF DDM_NAME = הצהרת שימוש בטבלה
- READ / FIND / STORE / UPDATE / DELETE = פעולות בסיס הנתונים
"""

PROGRAM_ANALYSIS_PROMPT = """
{system_context}

=== הקשר הקובץ ===
קובץ: {filename}
תוכנית: {program_name}
DDMs בשימוש בתוכנית זו (זוהו סטטית בקוד): {prog_ddms}
DDMs נוספים בקובץ — להקשר בלבד, אל תדווח עליהם כישויות של תוכנית זו: {other_ddms}
קורא ל (CALLNAT): {callnats}
תקצירי התוכניות הנקראות (מניתוח קודם — השתמש בהם להבנת הזרימה):
{callee_summaries}
נקרא מ: {called_by}

=== קוד התוכנית ===
{code}

=== משימה ===
נתח את הקוד והחזר JSON תקני בלבד (ללא markdown fences, ללא טקסט נוסף).
חשוב: דווח ב-entities רק על ישויות שתוכנית זו באמת משתמשת בהן —
כאלה שמופיעות ב-VIEW OF בקוד שלה או ששדותיהן נקראים/נכתבים בקוד בפועל.

{{
  "filename": "{filename}",
  "program_name": "{program_name}",
  "business_purpose": "מה התוכנית עושה — תיאור עסקי קצר",
  "complexity": "פשוט | בינוני | מורכב",

  "entities": [
    {{
      "name": "שם ה-DDM / ישות",
      "adabas_file": "מספר או שם קובץ ב-ADABAS",
      "description": "תיאור הישות",
      "fields": [
        {{
          "name": "שם שדה טכני",
          "label": "שם תצוגה בעברית",
          "type": "string | number | date | boolean | picklist | relation",
          "required": "כן | לא",
          "length": "אורך מקסימלי",
          "description": "תיאור השדה ומשמעותו"
        }}
      ]
    }}
  ],

  "workflows": [
    {{
      "name": "שם הזרימה / תהליך — צור זרימה נפרדת לכל: פעולת PF-key, סוג משתמש, תת-תהליך עיקרי, מסך",
      "trigger": "on_create | on_update | on_delete | on_field_change | scheduled | manual | pf_key",
      "trigger_details": "פירוט מדויק — לדוגמה: לחיצת PF3, בחירת אפשרות 2, כניסה ראשונה",
      "actors": "מי מעורב — תפקידים / משתמשים",
      "steps": [
        "צעד 1: תיאור מפורט",
        "צעד 2: IF תנאי → נתיב א / נתיב ב",
        "צעד 3: קריאה ל-CALLNAT XX / עדכון DDM YY"
      ],
      "business_rules": [
        {{
          "rule_id": "RULE-001",
          "name": "שם החוק",
          "logic": "IF תנאי THEN פעולה ELSE פעולה אחרת",
          "trigger": "מתי מופעל",
          "exceptions": "חריגות"
        }}
      ],
      "success_outcome": "תוצאה תקינה",
      "error_handling": "טיפול בשגיאות"
    }}
  ],

  "permissions": [
    {{
      "role": "שם תפקיד — לפי ערכי GL-MMAD-RASHAY-SODI או משתנה הרשאה אחר בקוד",
      "description": "תיאור התפקיד והרמה (לדוגמה: רמה < 6, רמה >= 8)",
      "entity": "שם הישות",
      "read": "all | own | none",
      "create": "כן | לא — לפי קוד STORE/INSERT",
      "update": "all | own | none — לפי קוד UPDATE",
      "delete": "all | own | none — לפי קוד DELETE"
    }}
  ],

  "integrations": [
    {{
      "system": "שם המערכת החיצונית בלבד — לא תוכניות Natural פנימיות",
      "type": "external | internal",
      "direction": "כניסה | יציאה | דו-כיווני",
      "data": "אילו נתונים עוברים",
      "frequency": "ריאל-טיים | אצווה | לפי דרישה",
      "notes": "הערות"
    }}
  ],

  "dependencies": [
    {{
      "type": "CALLNAT | DDM | ADABAS_FILE",
      "name": "שם",
      "purpose": "מטרת השימוש"
    }}
  ],

  "glossary": [
    {{"term": "מונח", "definition": "הגדרה"}}
  ],

  "open_questions": ["שאלה פתוחה 1"]
}}
"""

DDM_EXTRACT_PROMPT = """
{system_context}

=== קובץ: {filename} ===
{code}

=== משימה ===
חלץ את הגדרות ה-DDM המלאות מהקוד (DEFINE DATA, VIEW OF, שמות שדות, סוגים, אורכים).
החזר JSON תקני בלבד (ללא markdown fences):

{{
  "ddms": [
    {{
      "name": "שם ה-DDM",
      "adabas_file": "מספר/שם קובץ ADABAS",
      "description": "תיאור",
      "fields": [
        {{
          "name": "שם שדה",
          "natural_name": "שם בנטורל",
          "type": "A | N | D | T | P | L",
          "length": "אורך",
          "level": "רמה 1/2/3",
          "description": "תיאור"
        }}
      ]
    }}
  ]
}}
"""


SYNTHESIS_PROMPT = """
{system_context}

הישות (DDM) בשם {entity} זוהתה ב-{n_programs} תוכניות שונות.
להלן כל התצפיות שנאספו עליה מניתוחי התוכניות:

קובץ ADABAS (ערכים שנצפו): {adabas_files}

תיאורים שניתנו לישות:
{descriptions}

שדות שנצפו (שם שדה → ערכים שנצפו בתוכניות שונות):
{fields_block}

תוכניות שמשתמשות בישות: {programs}

=== משימה ===
מזג את כל התצפיות לרשומה קנונית אחת של הישות:
- בחר את התיאור המדויק והמלא ביותר (או נסח תיאור חדש שמאחד את כולם)
- אחד את רשימת השדות: כל שדה מופיע פעם אחת, עם הערכים הנכונים ביותר
- אם יש סתירות אמיתיות בין תוכניות (סוג שונה, אורך שונה לאותו שדה) — פרט אותן ב-conflicts
החזר JSON תקני בלבד:

{{
  "name": "{entity}",
  "adabas_file": "מספר/שם קובץ ADABAS",
  "description": "התיאור הקנוני הממוזג",
  "business_role": "תפקיד הישות במערכת כולה — משפט אחד",
  "fields": [
    {{"name": "שם שדה", "label": "תווית בעברית", "type": "string | number | date | boolean | picklist | relation",
      "length": "אורך", "required": "כן | לא", "description": "תיאור ממוזג"}}
  ],
  "conflicts": ["סתירה שדורשת בירור"]
}}
"""

OVERVIEW_PROMPT = """
{system_context}

להלן כל הישויות שזוהו במערכת לאחר איחוד, עם תיאורן ומספר התוכניות שמשתמשות בכל אחת:

{entities_block}

סטטיסטיקות: {n_programs} תוכניות נותחו, {n_entities} ישויות ייחודיות.

=== משימה ===
כתוב סקירת-על של המערכת כולה. החזר JSON תקני בלבד:

{{
  "system_purpose": "מה המערכת עושה — פסקה אחת",
  "domains": [
    {{"name": "שם תחום עסקי", "description": "תיאור התחום", "entities": ["הישויות השייכות לתחום"]}}
  ],
  "key_insights": ["תובנה מרכזית על המערכת"],
  "open_questions": ["שאלה שדורשת בירור עם בעלי הידע"]
}}
"""


def build_program_prompt(fd, prog, kb, ddm_cache, purpose_cache=None):
    """Build the analysis prompt for a single program."""
    filename     = fd['filename']
    prog_name    = prog['name']
    code         = prog['code'][:MAX_PROGRAM_CHARS]
    callnats     = prog.get('callnats', fd['callnats'])

    # Who calls this program?
    called_by = [f for f, calls in kb['call_graph'].items()
                 if prog_name.upper() in [c.upper() for c in calls]]

    # One-line summaries of callee programs analyzed in earlier waves
    callee_lines = []
    for cn in callnats:
        purpose = (purpose_cache or {}).get(cn)
        if purpose:
            callee_lines.append(f"  {cn}: {purpose[:200]}")

    prog_ddms  = prog.get('ddms', [])
    other_ddms = [d for d in fd['ddms'] if d not in prog_ddms]

    return PROGRAM_ANALYSIS_PROMPT.format(
        system_context   = SYSTEM_CONTEXT,
        filename         = filename,
        program_name     = prog_name,
        prog_ddms        = ', '.join(prog_ddms) or 'לא זוהו (ייתכן שימוש דרך LDA חיצוני)',
        other_ddms       = ', '.join(other_ddms) or 'אין',
        callnats         = ', '.join(callnats) or 'אין',
        callee_summaries = '\n'.join(callee_lines) or '  (אין מידע)',
        called_by        = ', '.join(called_by) or 'לא זוהה',
        code             = code,
    )


def data_analysis_from_ddms(fd, prog, raw):
    """Convert a DDM-extraction answer (data-only member) into the standard
    analysis shape so it flows into the same checkpoint/Excel/synthesis."""
    entities = []
    for d in (raw or {}).get('ddms', []):
        entities.append({
            'name'       : d.get('name', ''),
            'adabas_file': d.get('adabas_file', ''),
            'description': d.get('description', ''),
            'fields'     : [{'name': f.get('name', ''),
                             'label': f.get('natural_name', ''),
                             'type': f.get('type', ''),
                             'length': f.get('length', ''),
                             'required': '',
                             'description': f.get('description', '')}
                            for f in d.get('fields', [])],
        })
    return {'filename': fd['filename'], 'program_name': prog['name'],
            'program_type': 'data',
            'business_purpose': 'אזור נתונים (LDA/GDA/Copycode) — הגדרות בלבד, ללא לוגיקה',
            'complexity': 'פשוט', 'entities': entities, 'workflows': [],
            'permissions': [], 'integrations': [], 'dependencies': [],
            'glossary': [], 'open_questions': []}


# ──────────────────────────────────────────────────────────────────────────────
# SYNTHESIS (phase 3.5) — consolidate per-program analyses into one canonical
# record per entity + a system-level overview
# ──────────────────────────────────────────────────────────────────────────────

def collect_entity_observations(analyses):
    """Aggregate every appearance of each entity across all program analyses."""
    obs = {}
    for a in analyses:
        if 'error' in a:
            continue
        for e in a.get('entities', []):
            name = (e.get('name') or '').strip().upper()
            if not name:
                continue
            o = obs.setdefault(name, {'descriptions': [], 'adabas_files': [],
                                      'programs': [], 'fields': {}})
            desc = (e.get('description') or '').strip()
            if desc and desc not in o['descriptions'] and len(o['descriptions']) < 8:
                o['descriptions'].append(desc)
            af = str(e.get('adabas_file') or '').strip()
            if af and af not in o['adabas_files']:
                o['adabas_files'].append(af)
            o['programs'].append(a.get('program_name', ''))
            for f in e.get('fields', []):
                fn = (f.get('name') or '').strip().upper()
                if not fn:
                    continue
                fo = o['fields'].setdefault(fn, {'labels': [], 'types': [], 'lengths': [],
                                                 'requireds': [], 'descriptions': []})
                for src, dst, cap in (('label', 'labels', 4), ('type', 'types', 4),
                                      ('length', 'lengths', 4), ('required', 'requireds', 2),
                                      ('description', 'descriptions', 3)):
                    v = str(f.get(src) or '').strip()
                    if v and v not in fo[dst] and len(fo[dst]) < cap:
                        fo[dst].append(v)
    return obs


def build_synthesis_prompt(name, o):
    field_lines = []
    for fn, fo in list(o['fields'].items())[:400]:
        field_lines.append(f"  {fn}: תוויות={fo['labels']} סוגים={fo['types']} "
                           f"אורכים={fo['lengths']} חובה={fo['requireds']} "
                           f"תיאורים={fo['descriptions'][:2]}")
    descriptions = '\n'.join(f"  - {d[:300]}" for d in o['descriptions']) or '  (אין)'
    programs = sorted(set(o['programs']))
    progs_txt = ', '.join(programs[:40])
    if len(programs) > 40:
        progs_txt += f" ועוד {len(programs) - 40}"
    return SYNTHESIS_PROMPT.format(
        system_context = SYSTEM_CONTEXT,
        entity         = name,
        n_programs     = len(set(o['programs'])),
        adabas_files   = ', '.join(o['adabas_files']) or 'לא זוהה',
        descriptions   = descriptions,
        fields_block   = '\n'.join(field_lines) or '  (אין)',
        programs       = progs_txt,
    )


def static_merge_entity(name, o):
    """No-AI fallback: merge observations mechanically so the sheet is never empty."""
    fields = []
    for fn, fo in o['fields'].items():
        fields.append({
            'name'       : fn,
            'label'      : fo['labels'][0] if fo['labels'] else '',
            'type'       : fo['types'][0] if fo['types'] else '',
            'length'     : fo['lengths'][0] if fo['lengths'] else '',
            'required'   : fo['requireds'][0] if fo['requireds'] else '',
            'description': fo['descriptions'][0] if fo['descriptions'] else '',
        })
    return {
        'name'         : name,
        'adabas_file'  : o['adabas_files'][0] if o['adabas_files'] else '',
        'description'  : max(o['descriptions'], key=len, default=''),
        'business_role': '',
        'fields'       : sorted(fields, key=lambda f: f['name']),
        'conflicts'    : [],
        'consolidation': 'static',  # AI consolidation failed — mechanical merge
    }


def run_synthesis(api_key, analyses, workers):
    """Consolidate all entities + produce a system overview. Checkpointed:
    an entity is re-consolidated only if its observation count changed
    (e.g. failed programs were retried and added data)."""
    obs = collect_entity_observations(analyses)
    if not obs:
        return None

    n_programs = len([a for a in analyses if 'error' not in a])
    ckpt = load_checkpoint(SYNTHESIS_CHECKPOINT_PATH,
                           key_fn=lambda a: a.get('entity', ''))

    results = {}
    tasks   = []
    for name, o in sorted(obs.items()):
        n_obs = len(o['programs'])
        prev  = ckpt.get(name)
        if prev and 'error' not in prev and prev.get('n_obs') == n_obs:
            results[name] = prev['record']
        else:
            tasks.append((name, o, n_obs))

    print(f"\n-- Phase 3.5: synthesis — consolidating {len(obs)} entities --------")
    if results:
        print(f"  {len(results)} entities loaded from checkpoint")

    io_lock = threading.Lock()
    state   = {'completed': 0, 'errors': 0}

    def worker(name, o, n_obs):
        try:
            record, error = call_and_parse(api_key, build_synthesis_prompt(name, o))
        except Exception as e:
            record, error = None, str(e)
        with io_lock:
            state['completed'] += 1
            if error:
                state['errors'] += 1
                log_error(ERROR_LOG_PATH, 'SYNTHESIS', name, error)
                append_checkpoint(SYNTHESIS_CHECKPOINT_PATH,
                                  {'entity': name, 'n_obs': n_obs, 'error': error})
                results[name] = static_merge_entity(name, o)
                print(f"  WARN [{state['completed']}/{len(tasks)}] {name}  {error} — static merge", flush=True)
            else:
                record.setdefault('name', name)
                append_checkpoint(SYNTHESIS_CHECKPOINT_PATH,
                                  {'entity': name, 'n_obs': n_obs, 'record': record})
                results[name] = record
                print(f"  OK [{state['completed']}/{len(tasks)}] {name}  "
                      f"fields:{len(record.get('fields', []))}", flush=True)

    interrupted = False
    if tasks:
        executor = ThreadPoolExecutor(max_workers=workers)
        futures  = [executor.submit(worker, *t) for t in tasks]
        try:
            for fut in as_completed(futures):
                fut.result()
        except KeyboardInterrupt:
            interrupted = True
            print("\n  Synthesis interrupted — remaining entities merged statically; rerun to complete")
        finally:
            try:
                executor.shutdown(wait=True, cancel_futures=True)
            except KeyboardInterrupt:
                executor.shutdown(wait=False, cancel_futures=True)
        for name, o, n_obs in tasks:       # fill anything still missing
            if name not in results:
                results[name] = static_merge_entity(name, o)

    # usage info for the Excel sheet
    for name, o in obs.items():
        rec = results.get(name)
        if rec is not None:
            progs = sorted(set(o['programs']))
            rec['used_by_count'] = len(progs)
            rec['used_by']       = progs

    entities = [results[name] for name in sorted(results)]

    # ── System overview — one call over the consolidated picture ──────────────
    overview = None
    ov_prev  = ckpt.get('__OVERVIEW__')
    if not interrupted:
        if ov_prev and 'error' not in ov_prev and ov_prev.get('n_obs') == len(entities):
            overview = ov_prev['record']
        else:
            ent_lines = [f"  {e.get('name', '')}: "
                         f"{e.get('business_role') or e.get('description', '')[:150]} "
                         f"({e.get('used_by_count', 0)} תוכניות)"
                         for e in entities[:500]]
            prompt = OVERVIEW_PROMPT.format(system_context=SYSTEM_CONTEXT,
                                            entities_block='\n'.join(ent_lines),
                                            n_programs=n_programs,
                                            n_entities=len(entities))
            try:
                overview, ov_err = call_and_parse(api_key, prompt)
                if ov_err:
                    raise RuntimeError(ov_err)
                append_checkpoint(SYNTHESIS_CHECKPOINT_PATH,
                                  {'entity': '__OVERVIEW__', 'n_obs': len(entities),
                                   'record': overview})
                print("  System overview generated")
            except Exception as e:
                log_error(ERROR_LOG_PATH, 'SYNTHESIS', '__OVERVIEW__', str(e))
                print(f"  WARN system overview failed: {e}")

    synthesis = {'entities': entities, 'overview': overview}
    with open('synthesis.json', 'w', encoding='utf-8') as f:
        json.dump(synthesis, f, ensure_ascii=False, indent=2)
    print(f"  synthesis.json saved ({len(entities)} consolidated entities)")
    if state['errors']:
        print(f"  WARN {state['errors']} entities failed AI consolidation (merged statically) — see {ERROR_LOG_PATH}")
    return synthesis


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ──────────────────────────────────────────────────────────────────────────────

def add_header_row(ws, headers):
    ws.append(headers)
    for c in ws[ws.max_row]:
        hdr(c)
    freeze_header(ws)


def build_readme_sheet(wb):
    """Add a README sheet as the first sheet explaining all sheets and columns."""
    ws = wb.create_sheet("מדריך לקריאה", 0)

    TITLE_FILL  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    TITLE_FONT  = Font(color="FFFFFF", bold=True, name="Arial", size=13)
    SHEET_FILL  = PatternFill(start_color="2E5990", end_color="2E5990", fill_type="solid")
    SHEET_FONT  = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    COL_FILL    = PatternFill(start_color="D6E4F7", end_color="D6E4F7", fill_type="solid")
    COL_FONT    = Font(bold=True, name="Arial", size=10)
    BODY_FONT   = Font(name="Arial", size=10)
    THIN        = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'),  bottom=Side(style='thin'))
    WRAP_R      = Alignment(horizontal="right", vertical="top", wrap_text=True)
    WRAP_C      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def title_row(text, row):
        ws.cell(row=row, column=1, value=text).fill  = TITLE_FILL
        ws.cell(row=row, column=1).font              = TITLE_FONT
        ws.cell(row=row, column=1).alignment         = WRAP_C
        ws.cell(row=row, column=1).border            = THIN
        ws.cell(row=row, column=2, value="").fill    = TITLE_FILL
        ws.cell(row=row, column=2).border            = THIN
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    def sheet_row(text, row):
        ws.cell(row=row, column=1, value=text).fill  = SHEET_FILL
        ws.cell(row=row, column=1).font              = SHEET_FONT
        ws.cell(row=row, column=1).alignment         = WRAP_C
        ws.cell(row=row, column=1).border            = THIN
        ws.cell(row=row, column=2, value="").fill    = SHEET_FILL
        ws.cell(row=row, column=2).border            = THIN
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    def col_row(col_name, description, row, alt=False):
        c1 = ws.cell(row=row, column=1, value=col_name)
        c2 = ws.cell(row=row, column=2, value=description)
        c1.font  = COL_FONT
        c2.font  = BODY_FONT
        c1.alignment = WRAP_R
        c2.alignment = WRAP_R
        c1.border = THIN
        c2.border = THIN
        if alt:
            c1.fill = COL_FILL
            c2.fill = COL_FILL

    CONTENT = [
        ("מדריך — אפיון טכני Natural ADABAS", "title"),
        ("הקובץ נוצר אוטומטית על ידי סקריפט Python שמנתח קוד Natural ADABAS ומשתמש ב-Gemini AI.", "info"),
        ("כל שורה מייצגת תוכנית בודדת (member) שחולצה מתוך קבצי ה-.txt.", "info"),
        ("", "space"),

        ("גיליון 1 — סיכום", "sheet"),
        ("קובץ", "שם קובץ ה-.txt המקורי שממנו חולצה התוכנית"),
        ("תוכנית", "שם התוכנית כפי שמופיע בשורת ה-*C** בקובץ המיינפריים"),
        ("מטרה עסקית", "תיאור תמציתי של מה התוכנית עושה מבחינה עסקית"),
        ("DDMs", "רשימת ה-DDMs (טבלאות ADABAS) שבהן התוכנית משתמשת"),
        ("CALLNAT", "רשימת תוכניות חיצוניות שהתוכנית קוראת להן"),
        ("מורכבות", "הערכת מורכבות: פשוט / בינוני / מורכב"),
        ("", "space"),

        ("גיליון 2 — סקירת מערכת", "sheet"),
        ("נושא / תוכן", "תמונת-על של המערכת כולה: מטרה, תחומים עסקיים, תובנות ושאלות — מסונתז מכלל הניתוחים"),
        ("", "space"),

        ("גיליון 3 — ישויות מאוחדות", "sheet"),
        ("ישות / DDM", "שורה אחת לכל ישות במערכת כולה — התוצר המאוחד של כל הניתוחים"),
        ("תיאור קנוני", "תיאור אחד ממוזג מכל התוכניות שמשתמשות בישות"),
        ("תפקיד עסקי", "תפקיד הישות במערכת כולה"),
        ("מס' תוכניות / תוכניות", "כמה ואילו תוכניות משתמשות בישות"),
        ("קונפליקטים", "סתירות שהתגלו בין תוכניות (סוג/אורך שונים לאותו שדה) — דורש בירור"),
        ("אופן איחוד", "AI = איחוד מלא; מכני = ה-AI נכשל והמיזוג נעשה אוטומטית"),
        ("", "space"),

        ("גיליון 4 — שדות מאוחדים", "sheet"),
        ("ישות / שם שדה / ...", "רשימת השדות הסופית של כל ישות — כל שדה פעם אחת, ללא כפילויות בין תוכניות"),
        ("", "space"),

        ("גיליון 5 — ישויות (פר תוכנית)", "sheet"),
        ("קובץ", "שם קובץ המקור"),
        ("תוכנית", "שם התוכנית"),
        ("ישות / DDM", "שם ה-DDM — מקביל לטבלה בבסיס נתונים רלציוני"),
        ("קובץ ADABAS", "מספר או שם הקובץ הפיזי ב-ADABAS"),
        ("תיאור ישות", "הסבר עסקי של מה הישות מייצגת"),
        ("", "space"),

        ("גיליון 6 — שדות (פר תוכנית)", "sheet"),
        ("קובץ", "שם קובץ המקור"),
        ("תוכנית", "שם התוכנית"),
        ("ישות", "שם ה-DDM שאליו שייך השדה"),
        ("שם שדה", "שם השדה הטכני בקוד Natural"),
        ("תווית", "שם תצוגה בעברית"),
        ("סוג", "סוג הנתון: string | number | date | boolean | picklist | relation"),
        ("אורך", "אורך מקסימלי של השדה"),
        ("חובה", "האם השדה חובה: כן / לא"),
        ("תיאור", "הסבר עסקי של השדה ומשמעותו"),
        ("", "space"),

        ("גיליון 7 — זרימות עבודה", "sheet"),
        ("קובץ", "שם קובץ המקור"),
        ("תוכנית", "שם התוכנית"),
        ("שם זרימה", "שם התהליך העסקי — כל PF-key / פעולה / מסך הוא זרימה נפרדת"),
        ("טריגר", "סוג האירוע המפעיל: manual | on_create | on_update | pf_key | scheduled"),
        ("פירוט טריגר", "תיאור מפורט של מה גורם לתהליך להתחיל"),
        ("שחקנים", "מי מעורב: תפקידים / משתמשים / מערכות"),
        ("צעדים", "רצף הפעולות בתהליך, כולל הסתעפויות IF/THEN"),
        ("תוצאה תקינה", "מה קורה כשהתהליך מסתיים בהצלחה"),
        ("טיפול בשגיאות", "מה קורה כשיש שגיאה"),
        ("", "space"),

        ("גיליון 8 — חוקי עסק", "sheet"),
        ("קובץ", "שם קובץ המקור"),
        ("תוכנית", "שם התוכנית"),
        ("זרימה", "שם הזרימה שאליה שייך החוק"),
        ("RULE ID", "מזהה ייחודי של החוק: BR-<תוכנית>-001"),
        ("שם החוק", "שם קצר ומתאר של החוק העסקי"),
        ("לוגיקה", "הלוגיקה המדויקת: IF תנאי THEN פעולה ELSE פעולה"),
        ("טריגר", "מתי החוק מופעל"),
        ("חריגות", "מקרים שבהם החוק אינו חל"),
        ("", "space"),

        ("גיליון 9 — הרשאות ותפקידים", "sheet"),
        ("קובץ", "שם קובץ המקור"),
        ("תוכנית", "שם התוכנית"),
        ("תפקיד", "שם התפקיד — נגזר מבדיקות הרשאה בקוד (לדוגמה: GL-MMAD-RASHAY-SODI)"),
        ("תיאור תפקיד", "הסבר מה התפקיד כולל ומה רמת ההרשאה שלו"),
        ("ישות", "הישות (DDM) שעליה חלה ההרשאה"),
        ("קריאה", "הרשאת קריאה: all | own | none"),
        ("יצירה", "הרשאת יצירה: כן / לא"),
        ("עדכון", "הרשאת עדכון: all | own | none"),
        ("מחיקה", "הרשאת מחיקה: all | own | none"),
        ("", "space"),

        ("גיליון 10 — אינטגרציות חיצוניות", "sheet"),
        ("קובץ", "שם קובץ המקור"),
        ("תוכנית", "שם התוכנית"),
        ("מערכת חיצונית", "שם המערכת החיצונית — לא כולל תוכניות Natural פנימיות"),
        ("כיוון", "כיוון הנתונים: כניסה | יציאה | דו-כיווני"),
        ("נתונים", "אילו נתונים עוברים בין המערכות"),
        ("תדירות", "ריאל-טיים | אצווה | לפי דרישה"),
        ("הערות", "הערות נוספות על האינטגרציה"),
        ("", "space"),

        ("גיליון 11 — גרף קריאות", "sheet"),
        ("קובץ מקור", "קובץ ה-.txt שממנו מגיעה הקריאה"),
        ("תוכנית מקור", "שם התוכנית שמבצעת את הקריאה"),
        ("סוג", "סוג התלות: CALLNAT (קריאה לתוכנית) | DDM (שימוש בטבלה) | ADABAS_FILE"),
        ("קובץ/תוכנית יעד", "שם התוכנית או ה-DDM שאליו מתבצעת הקריאה"),
        ("מטרה", "הסבר של מטרת הקריאה"),
        ("", "space"),

        ("גיליון 12 — גלוסרי", "sheet"),
        ("מונח", "המונח הטכני (לרוב באנגלית/עברית מקוצרת) כפי שמופיע בקוד"),
        ("הגדרה", "הסבר בעברית של משמעות המונח"),
        ("קובץ", "הקובץ שממנו חולץ המונח"),
        ("", "space"),

        ("גיליון 13 — שאלות פתוחות", "sheet"),
        ("קובץ", "שם קובץ המקור"),
        ("תוכנית", "שם התוכנית"),
        ("שאלה", "שאלה שעלתה בניתוח ודורשת בירור עם בעל הידע"),
        ("", "space"),

        ("גיליון 14 — ולידציה", "sheet"),
        ("קובץ / תוכנית", "היכן נמצא הפער"),
        ("סוג", "ישות (VIEW OF) או CALLNAT"),
        ("שם", "שם הישות / התוכנית שבמחלוקת"),
        ("ממצא", "הצלבה אוטומטית בין סריקת הקוד לדיווח ה-AI. שלוש רמות: חשד להזיה = שם שלא קיים בשום מקום במערכת (חמור); שיוך שגוי = ישות/תוכנית אמיתית שיוחסה לתוכנית הלא נכונה (קל — לרוב LDA חיצוני); פספוס = נמצא בקוד ולא דווח"),
    ]

    row = 1
    alt = False
    for item in CONTENT:
        if item[1] == "title":
            title_row(item[0], row)
            ws.row_dimensions[row].height = 28
        elif item[1] == "sheet":
            sheet_row(item[0], row)
            ws.row_dimensions[row].height = 22
            alt = False
        elif item[1] == "info":
            c = ws.cell(row=row, column=1, value=item[0])
            c.font = BODY_FONT
            c.alignment = WRAP_R
            c.border = THIN
            ws.cell(row=row, column=2).border = THIN
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.row_dimensions[row].height = 18
        elif item[1] == "space":
            ws.row_dimensions[row].height = 8
        else:
            col_row(item[0], item[1], row, alt)
            ws.row_dimensions[row].height = 30
            alt = not alt
        row += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 70
    ws.sheet_view.rightToLeft = True


def build_synthesis_sheets(wb, synthesis):
    """Sheets 2-4: system overview + consolidated entities/fields (phase 3.5)."""
    overview = synthesis.get('overview')
    entities = synthesis.get('entities') or []

    # ── 2. System overview ───────────────────────────────────────────────────
    ws = wb.create_sheet("סקירת מערכת", 2)
    add_header_row(ws, ["נושא", "תוכן"])
    rows = []
    if overview:
        rows.append(("מטרת המערכת", overview.get('system_purpose', '')))
        for d in overview.get('domains', []):
            ents = ', '.join(d.get('entities', []))
            rows.append((f"תחום: {d.get('name', '')}",
                         f"{d.get('description', '')}\nישויות: {ents}"))
        for ins in overview.get('key_insights', []):
            rows.append(("תובנה", ins))
        for q in overview.get('open_questions', []):
            rows.append(("שאלה פתוחה", q))
    else:
        rows.append(("—", "הסקירה לא נוצרה בהרצה זו — הרץ שוב להשלמה"))
    for i, (k, v) in enumerate(rows):
        ws.append([k, v])
        for c in ws[ws.max_row]:
            cell_style(c, alt=(i % 2 == 0))
    set_col_widths(ws, {'A': 25, 'B': 100})
    ws.sheet_view.rightToLeft = True

    # ── 3. Consolidated entities ─────────────────────────────────────────────
    ws = wb.create_sheet("ישויות מאוחדות", 3)
    add_header_row(ws, ["ישות / DDM", "קובץ ADABAS", "תיאור קנוני", "תפקיד עסקי",
                        "מס' תוכניות", "תוכניות", "קונפליקטים", "אופן איחוד"])
    for i, e in enumerate(entities):
        progs = e.get('used_by', [])
        progs_txt = ', '.join(progs[:30])
        if len(progs) > 30:
            progs_txt += f" ועוד {len(progs) - 30}"
        ws.append([
            e.get('name', ''), e.get('adabas_file', ''), e.get('description', ''),
            e.get('business_role', ''), e.get('used_by_count', ''),
            progs_txt, '\n'.join(e.get('conflicts') or []),
            'מכני' if e.get('consolidation') == 'static' else 'AI',
        ])
        for c in ws[ws.max_row]:
            cell_style(c, alt=(i % 2 == 0))
    set_col_widths(ws, {'A':22,'B':14,'C':55,'D':40,'E':12,'F':45,'G':40,'H':12})

    # ── 4. Consolidated fields ───────────────────────────────────────────────
    ws = wb.create_sheet("שדות מאוחדים", 4)
    add_header_row(ws, ["ישות", "שם שדה", "תווית", "סוג", "אורך", "חובה", "תיאור"])
    i = 0
    for e in entities:
        for f in e.get('fields', []):
            ws.append([e.get('name', ''), f.get('name', ''), f.get('label', ''),
                       f.get('type', ''), f.get('length', ''), f.get('required', ''),
                       f.get('description', '')])
            for c in ws[ws.max_row]:
                cell_style(c, alt=(i % 2 == 0))
            i += 1
    set_col_widths(ws, {'A':22,'B':22,'C':22,'D':12,'E':10,'F':10,'G':55})


def build_excel(analyses, kb, output_path, synthesis=None, validation_rows=None):
    wb = openpyxl.Workbook()

    # ── 1. Summary ────────────────────────────────────────────────────────────
    # Grab the default sheet BEFORE inserting the README at index 0 —
    # otherwise wb.active points at the README and the summary lands inside it
    ws1 = wb.active
    ws1.title = "סיכום"
    build_readme_sheet(wb)
    add_header_row(ws1, ["קובץ", "תוכנית", "מטרה עסקית", "DDMs", "CALLNAT", "מורכבות"])

    for i, a in enumerate(analyses):
        if 'error' in a:
            ws1.append([a.get('filename', ''), a.get('program_name', ''), f"⚠ שגיאה: {a['error']}", '', '', ''])
        else:
            ents  = [e['name'] for e in a.get('entities', [])]
            deps  = [d['name'] for d in a.get('dependencies', []) if d.get('type') == 'CALLNAT']
            ws1.append([
                a.get('filename', ''),
                a.get('program_name', ''),
                a.get('business_purpose', ''),
                ', '.join(ents),
                ', '.join(deps),
                a.get('complexity', ''),
            ])
        for c in ws1[ws1.max_row]:
            cell_style(c, alt=(i % 2 == 0))

    set_col_widths(ws1, {'A':25,'B':25,'C':55,'D':30,'E':30,'F':15})

    if synthesis:
        build_synthesis_sheets(wb, synthesis)

    # ── 2. Entities ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("ישויות")
    add_header_row(ws2, ["קובץ", "תוכנית", "ישות / DDM", "קובץ ADABAS", "תיאור ישות"])

    i = 0
    for a in analyses:
        for e in a.get('entities', []):
            ws2.append([
                a.get('filename', ''), a.get('program_name', ''),
                e.get('name', ''), e.get('adabas_file', ''), e.get('description', ''),
            ])
            for c in ws2[ws2.max_row]: cell_style(c, alt=(i % 2 == 0))
            i += 1

    set_col_widths(ws2, {'A':25,'B':25,'C':25,'D':15,'E':55})

    # ── 3. Fields ─────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("שדות")
    add_header_row(ws3, ["קובץ", "תוכנית", "ישות", "שם שדה", "תווית", "סוג", "אורך", "חובה", "תיאור"])

    i = 0
    for a in analyses:
        for e in a.get('entities', []):
            entity_name = e.get('name', '')
            for f in e.get('fields', []):
                ws3.append([
                    a.get('filename', ''), a.get('program_name', ''), entity_name,
                    f.get('name', ''), f.get('label', ''), f.get('type', ''),
                    f.get('length', ''), f.get('required', ''), f.get('description', ''),
                ])
                for c in ws3[ws3.max_row]: cell_style(c, alt=(i % 2 == 0))
                i += 1

    set_col_widths(ws3, {'A':20,'B':20,'C':20,'D':20,'E':20,'F':12,'G':10,'H':10,'I':45})

    # ── 4. Workflows ──────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("זרימות עבודה")
    add_header_row(ws4, ["קובץ", "תוכנית", "שם זרימה", "טריגר", "פירוט טריגר", "שחקנים", "צעדים", "תוצאה תקינה", "טיפול בשגיאות"])

    i = 0
    for a in analyses:
        for wf in a.get('workflows', []):
            steps_text = '\n'.join(wf.get('steps', []))
            ws4.append([
                a.get('filename', ''), a.get('program_name', ''),
                wf.get('name', ''), wf.get('trigger', ''), wf.get('trigger_details', ''),
                wf.get('actors', ''), steps_text,
                wf.get('success_outcome', ''), wf.get('error_handling', ''),
            ])
            for c in ws4[ws4.max_row]: cell_style(c, alt=(i % 2 == 0))
            i += 1

    set_col_widths(ws4, {'A':20,'B':20,'C':30,'D':18,'E':35,'F':25,'G':60,'H':40,'I':40})

    # ── 5. Business Rules ─────────────────────────────────────────────────────
    ws5 = wb.create_sheet("חוקי עסק")
    add_header_row(ws5, ["קובץ", "תוכנית", "זרימה", "RULE ID", "שם החוק", "לוגיקה", "טריגר", "חריגות"])

    i = 0
    for a in analyses:
        for wf in a.get('workflows', []):
            wf_name = wf.get('name', '')
            for rule in wf.get('business_rules', []):
                ws5.append([
                    a.get('filename', ''), a.get('program_name', ''), wf_name,
                    rule.get('rule_id', ''), rule.get('name', ''),
                    rule.get('logic', ''), rule.get('trigger', ''), rule.get('exceptions', ''),
                ])
                for c in ws5[ws5.max_row]: cell_style(c, alt=(i % 2 == 0))
                i += 1

    set_col_widths(ws5, {'A':20,'B':20,'C':25,'D':12,'E':30,'F':60,'G':30,'H':30})

    # ── 6. Permissions ────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("הרשאות ותפקידים")
    add_header_row(ws6, ["קובץ", "תוכנית", "תפקיד", "תיאור תפקיד", "ישות", "קריאה", "יצירה", "עדכון", "מחיקה"])

    i = 0
    for a in analyses:
        for p in a.get('permissions', []):
            ws6.append([
                a.get('filename', ''), a.get('program_name', ''),
                p.get('role', ''), p.get('description', ''), p.get('entity', ''),
                p.get('read', ''), p.get('create', ''),
                p.get('update', ''), p.get('delete', ''),
            ])
            for c in ws6[ws6.max_row]: cell_style(c, alt=(i % 2 == 0))
            i += 1

    set_col_widths(ws6, {'A':20,'B':20,'C':25,'D':40,'E':20,'F':12,'G':12,'H':12,'I':12})

    # ── 7. External Integrations only ────────────────────────────────────────
    ws7 = wb.create_sheet("אינטגרציות חיצוניות")
    add_header_row(ws7, ["קובץ", "תוכנית", "מערכת חיצונית", "כיוון", "נתונים", "תדירות", "הערות"])

    i = 0
    for a in analyses:
        for intg in a.get('integrations', []):
            if intg.get('type', '').lower() == 'internal':
                continue  # internal CALLNAT calls go to call graph sheet
            ws7.append([
                a.get('filename', ''), a.get('program_name', ''),
                intg.get('system', ''), intg.get('direction', ''),
                intg.get('data', ''), intg.get('frequency', ''), intg.get('notes', ''),
            ])
            for c in ws7[ws7.max_row]: cell_style(c, alt=(i % 2 == 0))
            i += 1

    set_col_widths(ws7, {'A':20,'B':20,'C':30,'D':15,'E':50,'F':18,'G':40})

    set_col_widths(ws7, {'A':20,'B':20,'C':25,'D':15,'E':50,'F':18,'G':40})

    # ── 8. Call Graph ─────────────────────────────────────────────────────────
    ws8 = wb.create_sheet("גרף קריאות")
    add_header_row(ws8, ["קובץ מקור", "תוכנית מקור", "סוג", "קובץ/תוכנית יעד", "מטרה"])

    i = 0
    for a in analyses:
        for dep in a.get('dependencies', []):
            ws8.append([
                a.get('filename', ''), a.get('program_name', ''),
                dep.get('type', ''), dep.get('name', ''), dep.get('purpose', ''),
            ])
            for c in ws8[ws8.max_row]: cell_style(c, alt=(i % 2 == 0))
            i += 1

    set_col_widths(ws8, {'A':25,'B':25,'C':15,'D':25,'E':50})

    # ── 9. Glossary ───────────────────────────────────────────────────────────
    ws9 = wb.create_sheet("גלוסרי")
    add_header_row(ws9, ["מונח", "הגדרה", "קובץ"])

    seen  = set()
    i = 0
    for a in analyses:
        for g in a.get('glossary', []):
            term = g.get('term', '')
            if term and term not in seen:
                seen.add(term)
                ws9.append([term, g.get('definition', ''), a.get('filename', '')])
                for c in ws9[ws9.max_row]: cell_style(c, alt=(i % 2 == 0))
                i += 1

    set_col_widths(ws9, {'A':30,'B':65,'C':25})

    # ── 10. Open Questions ────────────────────────────────────────────────────
    ws10 = wb.create_sheet("שאלות פתוחות")
    add_header_row(ws10, ["קובץ", "תוכנית", "שאלה"])

    i = 0
    for a in analyses:
        for q in a.get('open_questions', []):
            ws10.append([a.get('filename', ''), a.get('program_name', ''), q])
            for c in ws10[ws10.max_row]: cell_style(c, alt=(i % 2 == 0))
            i += 1

    set_col_widths(ws10, {'A':25,'B':25,'C':80})

    # ── 11. Validation ────────────────────────────────────────────────────────
    ws11 = wb.create_sheet("ולידציה")
    add_header_row(ws11, ["קובץ", "תוכנית", "סוג", "שם", "ממצא"])

    if validation_rows:
        for i, row in enumerate(validation_rows):
            ws11.append(list(row))
            for c in ws11[ws11.max_row]: cell_style(c, alt=(i % 2 == 0))
    else:
        ws11.append(["—", "—", "—", "—", "לא נמצאו פערים — דיווחי ה-AI עקביים עם הקוד"])
        for c in ws11[ws11.max_row]: cell_style(c)

    set_col_widths(ws11, {'A':20,'B':20,'C':12,'D':25,'E':70})

    wb.save(output_path)


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    # ── Debug mode ─────────────────────────────────────────────────────────────
    if len(sys.argv) >= 2 and sys.argv[1] == '--debug':
        target = sys.argv[2] if len(sys.argv) >= 3 else None
        if not target:
            # Find first .txt file
            txts = sorted(Path('.').glob('*.txt'))
            if not txts:
                print("No .txt files found.")
                sys.exit(1)
            target = str(txts[0])
        debug_file(target)
        return  # debug_file calls sys.exit, but just in case

    # ── Review/pack modes — quality-check from existing outputs, no API ───────
    for flag, fn in (('--review', run_review), ('--pack', run_pack)):
        if flag in sys.argv:
            idx = sys.argv.index(flag)
            size = 8
            if len(sys.argv) > idx + 1 and sys.argv[idx + 1].isdigit():
                size = max(1, int(sys.argv[idx + 1]))
            fn(size)
            return

    # ── Parallelism level ──────────────────────────────────────────────────────
    workers = MAX_WORKERS
    if '--workers' in sys.argv:
        try:
            workers = max(1, int(sys.argv[sys.argv.index('--workers') + 1]))
        except (IndexError, ValueError):
            print("Usage: python natural_adabas_spec_generator.py --workers N")
            sys.exit(1)

    print("=" * 62)
    print("  Natural ADABAS -> Technical Spec Generator  ")
    print("=" * 62)

    # ── Get API key ────────────────────────────────────────────────────────────
    api_key = os.environ.get('GEMINI_API_KEY', '').strip()
    if not api_key:
        api_key = input("\nEnter Gemini API key: ").strip()
    if not api_key:
        print("Error: missing API key. Aborting.")
        sys.exit(1)

    # ── Find .txt files ────────────────────────────────────────────────────────
    txt_files = sorted(Path('.').glob('*.txt'))
    if not txt_files:
        print("No .txt files found in the current directory.")
        sys.exit(1)

    print(f"\nFound {len(txt_files)} files:")
    for fp in txt_files:
        print(f"  {fp.name:35s}  {fp.stat().st_size/1024:.0f} KB")

    # ── PHASE 1: Static parse ──────────────────────────────────────────────────
    print("\n-- Phase 1: static parse ------------------------------------")
    files_data = []
    total_programs = 0
    for fp in txt_files:
        fd = parse_file(fp)
        files_data.append(fd)
        n_prog = len(fd['programs'])
        total_programs += n_prog
        print(f"  {fd['filename']:35s} "
              f"programs: {n_prog:3d}  DDMs: {len(fd['ddms']):3d}  "
              f"CALLNATs: {len(fd['callnats'])}")

    print(f"\n  Total programs/routines: {total_programs}")

    # ── Resume: load checkpoint from a previous (interrupted) run ─────────────
    # An entry is reused only if the program's current code hash matches the
    # one stored at analysis time — so parsing fixes (e.g. the *S** source
    # prefix) automatically invalidate analyses made on misread code.
    checkpoint = load_checkpoint(CHECKPOINT_PATH)
    prog_hash  = {(fd['filename'], p['name']): p['hash']
                  for fd in files_data for p in fd['programs']}
    done, stale = set(), set()
    for k, a in checkpoint.items():
        if 'error' in a:
            continue
        h = a.get('code_hash')
        if h is not None and k in prog_hash and h != prog_hash[k]:
            stale.add(k)
        else:
            done.add(k)
    remaining = [(fd, prog) for fd in files_data for prog in fd['programs']
                 if (fd['filename'], prog['name']) not in done]
    if done or stale:
        print(f"\n  Checkpoint found ({CHECKPOINT_PATH}): "
              f"{len(done)} programs already analyzed, {len(remaining)} remaining")
        print(f"    (delete the file to re-analyze everything)")
    if stale:
        print(f"  {len(stale)} analyses invalidated — code parsing changed since "
              f"they were analyzed (e.g. *S** prefix) — will be redone")

    # ── Token estimate warning (remaining unique programs only) ───────────────
    if remaining:
        ckpt_hashes = {a.get('code_hash') for a in checkpoint.values()
                       if 'error' not in a and a.get('code_hash')}
        uniq, seen_h = [], set(ckpt_hashes) - {None}
        for _, p in remaining:
            if p['hash'] not in seen_h:
                seen_h.add(p['hash'])
                uniq.append(p)
        n_dups = len(remaining) - len(uniq)
        if n_dups:
            print(f"\n  {n_dups} duplicates detected (identical code) — analyzed only once")
        total_chars   = sum(len(p['code']) for p in uniq)
        est_in_tokens = total_chars // 4          # ~4 chars per token
        est_out_tok   = len(uniq) * 3000          # ~3K output tokens per program
        print(f"\n  Token estimate ({len(uniq)} unique programs):")
        print(f"     input:  ~{est_in_tokens:,} tokens  ({total_chars/1e6:.1f}MB)")
        print(f"     output: ~{est_out_tok:,} tokens")
        print(f"     total:  ~{(est_in_tokens+est_out_tok):,} tokens")
        ans = input("  Continue? (y/n): ").strip().lower()
        if ans != 'y':
            print("  Cancelled.")
            sys.exit(0)
    else:
        print("\n  All programs already analyzed — skipping straight to output build")

    # ── PHASE 2: Knowledge base ────────────────────────────────────────────────
    print("\n-- Phase 2: knowledge base ----------------------------------")
    kb = build_knowledge_base(files_data)
    print(f"  Unique DDMs: {len(kb['all_ddms'])}")
    print(f"  CALLNAT links: {sum(len(v) for v in kb['call_graph'].values())}")

    kb_path = Path('knowledge_base.json')
    with open(kb_path, 'w', encoding='utf-8') as f:
        json.dump(kb, f, ensure_ascii=False, indent=2)
    print(f"  {kb_path} saved")

    # ── PHASE 3: Gemini analysis (parallel) ────────────────────────────────────
    print(f"\n-- Phase 3: AI analysis ({len(remaining)}/{total_programs} programs, "
          f"{workers} parallel calls) ----------")

    # Ordered task list; checkpointed programs are pre-filled into results
    order   = []   # (seq, key) in original program order — keeps Excel stable
    tasks   = []   # (seq, fd, prog) still needing analysis
    results = {}   # seq → analysis dict
    seq = 0
    for fd in files_data:
        for prog in fd['programs']:
            seq += 1
            key = (fd['filename'], prog['name'])
            order.append((seq, key))
            if key in done:
                results[seq] = checkpoint[key]
            else:
                tasks.append((seq, fd, prog))
    skipped = len(order) - len(tasks)

    # ── Duplicate detection: identical members exported in several files are
    #    analyzed once; the copies reuse the result without an API call ───────
    hash_to_analysis = {}   # code hash → successful analysis (checkpoint + this run)
    for a in checkpoint.values():
        if 'error' not in a and a.get('code_hash'):
            hash_to_analysis.setdefault(a['code_hash'], a)

    primary_tasks, dup_tasks = [], []
    claimed_hashes = set(hash_to_analysis)
    for t in tasks:
        h = t[2]['hash']
        if h in claimed_hashes:
            dup_tasks.append(t)
        else:
            claimed_hashes.add(h)
            primary_tasks.append(t)
    if dup_tasks:
        print(f"  {len(dup_tasks)} duplicates — analyzed once, copied at the end")

    n_data = sum(1 for _, _, p in primary_tasks if p.get('ptype') == 'data')
    if n_data:
        print(f"  Classification: {len(primary_tasks) - n_data} logic programs, "
              f"{n_data} data-only members (short prompt)")

    # Rebuild the DDM cache and the callee-purpose cache from checkpointed
    # analyses (valid ones only — not entries invalidated by a hash change)
    ddm_cache     = {}
    purpose_cache = {}   # PROG NAME → business_purpose, feeds caller prompts
    for k, a in checkpoint.items():
        if k in done:
            for e in a.get('entities', []):
                ddm_cache[e['name']] = e
            if a.get('business_purpose'):
                purpose_cache[(a.get('program_name') or '').upper()] = a['business_purpose']

    # ── Wave ordering: callees before callers, so caller prompts can include
    #    one-line summaries of the subprograms they CALLNAT ───────────────────
    name_to_tasks = defaultdict(list)
    for t in primary_tasks:
        name_to_tasks[t[2]['name'].upper()].append(t)
    pending  = set(name_to_tasks)
    calls_of = {}
    for n, ts in name_to_tasks.items():
        cs = set()
        for t in ts:
            cs |= {c for c in t[2].get('callnats', []) if c in pending and c != n}
        calls_of[n] = cs

    waves, assigned, left = [], set(), set(pending)
    while left:
        ready = {n for n in left if calls_of[n] <= assigned}
        if not ready:           # call cycle — analyze the rest together
            ready = set(left)
        waves.append(sorted(ready))
        assigned |= ready
        left     -= ready

    io_lock = threading.Lock()
    state   = {'completed': 0, 'errors': 0, 'start': time.time()}

    def handle_result(seq_no, fd, prog, analysis, error):
        """Runs in the worker thread — checkpoint, log and report under one lock,
        so results survive even if the main thread was interrupted meanwhile."""
        with io_lock:
            state['completed'] += 1
            n_done, n_total = state['completed'], len(primary_tasks)
            elapsed = time.time() - state['start']
            eta_min = (elapsed / n_done) * (n_total - n_done) / 60
            label   = f"[{n_done}/{n_total}] {fd['filename']} › {prog['name']}"
            if error:
                state['errors'] += 1
                entry = {'filename': fd['filename'], 'program_name': prog['name'],
                         'error': error}
                log_error(ERROR_LOG_PATH, fd['filename'], prog['name'], error)
                print(f"  WARN {label}  {error}", flush=True)
            else:
                entry = analysis
                # Force the identity fields — the checkpoint key depends on
                # them, and the model can't be trusted to echo them verbatim
                entry['filename']     = fd['filename']
                entry['program_name'] = prog['name']
                entry['code_hash']    = prog['hash']
                hash_to_analysis.setdefault(prog['hash'], entry)
                for e in analysis.get('entities', []):
                    ddm_cache[e['name']] = e
                if analysis.get('business_purpose'):
                    purpose_cache[prog['name'].upper()] = analysis['business_purpose']
                n_ent = len(analysis.get('entities', []))
                n_wf  = len(analysis.get('workflows', []))
                n_req = sum(len(wf.get('business_rules', [])) for wf in analysis.get('workflows', []))
                print(f"  OK {label}  entities:{n_ent} flows:{n_wf} rules:{n_req}  "
                      f"~{eta_min/60:.1f}h left", flush=True)
            append_checkpoint(CHECKPOINT_PATH, entry)
            results[seq_no] = entry

    def worker(seq_no, fd, prog):
        try:
            if prog.get('ptype') == 'data':
                # data-only member — short field-extraction prompt
                prompt = DDM_EXTRACT_PROMPT.format(system_context=SYSTEM_CONTEXT,
                                                   filename=fd['filename'],
                                                   code=prog['code'][:MAX_PROGRAM_CHARS])
                raw, error = call_and_parse(api_key, prompt)
                analysis   = data_analysis_from_ddms(fd, prog, raw) if not error else None
            else:
                prompt = build_program_prompt(fd, prog, kb, ddm_cache, purpose_cache)
                analysis, error = call_and_parse(api_key, prompt)
                if analysis is not None:
                    analysis['program_type'] = 'logic'
        except Exception as e:
            analysis, error = None, str(e)
        handle_result(seq_no, fd, prog, analysis, error)

    interrupted = False
    if tasks:
        if len(waves) > 1:
            print(f"  Call-graph wave order: {len(waves)} waves "
                  f"(subprograms first, then their callers)")
        executor = ThreadPoolExecutor(max_workers=workers)
        futures  = []
        try:
            for w_idx, wave_names in enumerate(waves):
                wave_tasks = [t for n in wave_names for t in name_to_tasks[n]]
                if len(waves) > 1:
                    print(f"  -- wave {w_idx + 1}/{len(waves)} ({len(wave_tasks)} programs) --")
                futures = [executor.submit(worker, *t) for t in wave_tasks]
                for fut in as_completed(futures):
                    fut.result()
        except KeyboardInterrupt:
            interrupted = True
            print(f"\n  Interrupted — waiting for in-flight calls to finish (up to a few minutes)...")
            print(f"    Everything completed is checkpointed; rerun to resume from this point.")
        finally:
            try:
                executor.shutdown(wait=True, cancel_futures=True)
            except KeyboardInterrupt:
                interrupted = True
                executor.shutdown(wait=False, cancel_futures=True)

    # ── Copy results onto duplicate programs (no API calls) ───────────────────
    if dup_tasks and not interrupted:
        copied = 0
        for seq_no, fd, prog in dup_tasks:
            src = hash_to_analysis.get(prog['hash'])
            if src is not None:
                entry = json.loads(json.dumps(src))   # deep copy
                entry['filename']     = fd['filename']
                entry['program_name'] = prog['name']
                entry['duplicate_of'] = f"{src.get('program_name', '')} ({src.get('filename', '')})"
                copied += 1
            else:   # the identical primary failed — both retried next run
                entry = {'filename': fd['filename'], 'program_name': prog['name'],
                         'error': 'Duplicate — analysis of the identical program failed; will retry next run'}
                state['errors'] += 1
            append_checkpoint(CHECKPOINT_PATH, entry)
            results[seq_no] = entry
        if copied:
            print(f"\n  {copied} duplicates copied from existing analyses (no API calls)")

    errors = state['errors']
    if skipped:
        print(f"\n  {skipped} programs loaded from checkpoint (not re-sent to API)")
    if errors:
        print(f"  WARN {errors} programs failed — details in {ERROR_LOG_PATH}")

    analyses = [results[s] for s, _ in order if s in results]

    # Save raw analyses
    analyses_path = Path('analyses.json')
    with open(analyses_path, 'w', encoding='utf-8') as f:
        json.dump(analyses, f, ensure_ascii=False, indent=2)
    print(f"\n  {analyses_path} saved")

    # ── PHASE 3.5: Synthesis — consolidate entities + system overview ─────────
    synthesis = None
    if interrupted:
        print("\n  (synthesis will run on the next run, after all programs are analyzed)")
    elif '--no-synthesis' in sys.argv:
        print("\n  (synthesis skipped — --no-synthesis)")
    else:
        synthesis = run_synthesis(api_key, analyses, workers)

    # ── PHASE 3.6: Validation — static cross-check, no API calls ──────────────
    validation_rows = build_validation(files_data, analyses)
    if validation_rows:
        n_hall = len([r for r in validation_rows if 'חשד להזיה' in r[4]])
        n_attr = len([r for r in validation_rows if 'שיוך שגוי' in r[4]])
        n_miss = len([r for r in validation_rows if 'פספוס' in r[4]])
        print(f"\n  Validation: {len(validation_rows)} gaps between code and AI reports "
              f"(hallucination suspicions: {n_hall}, misattributions: {n_attr}, "
              f"misses: {n_miss}) — see the validation sheet")
    else:
        print(f"\n  Validation: AI reports consistent with the code — no gaps")

    # ── PHASE 4: Build Excel ───────────────────────────────────────────────────
    print("\n-- Phase 4: building Excel ----------------------------------")
    output_path = Path('natural_adabas_spec.xlsx')
    build_excel(analyses, kb, output_path, synthesis, validation_rows)

    # ── Summary ────────────────────────────────────────────────────────────────
    print("\n" + "=" * 62)
    if interrupted:
        print("  Interrupted — partial output created")
    else:
        print("  Done!")
    print(f"  {output_path}")
    print(f"  {kb_path}")
    print(f"  {analyses_path}")
    if synthesis:
        print(f"  synthesis.json")
    if errors:
        print(f"  WARN {errors} programs failed — details in {ERROR_LOG_PATH}")
    if interrupted or errors:
        print(f"  Rerun the script to resume from this point "
              f"(completed programs are not re-sent)")
    elif CHECKPOINT_PATH.exists():
        print(f"  Checkpoint kept at {CHECKPOINT_PATH} — delete it to re-analyze everything")

    # Print sheet summary
    successful = [a for a in analyses if 'error' not in a]
    total_entities  = sum(len(a.get('entities', []))    for a in successful)
    total_fields    = sum(len(f) for a in successful
                          for e in a.get('entities', [])
                          for f in [e.get('fields', [])])
    total_workflows = sum(len(a.get('workflows', []))   for a in successful)
    total_rules     = sum(len(wf.get('business_rules', []))
                          for a in successful for wf in a.get('workflows', []))
    total_perms     = sum(len(a.get('permissions', [])) for a in successful)

    print(f"\n  Statistics:")
    print(f"    programs analyzed: {len(successful):>5}")
    print(f"    entities (DDMs):   {total_entities:>5}")
    print(f"    fields:            {total_fields:>5}")
    print(f"    workflows:         {total_workflows:>5}")
    print(f"    business rules:    {total_rules:>5}")
    print(f"    permissions:       {total_perms:>5}")
    print("=" * 62)


if __name__ == "__main__":
    main()
